"""
parent/bulk_views.py

Views for:
  1. POST   /parents/bulk-upload/                        — accept file, enqueue Celery task
  2. GET    /parents/bulk-upload/<id>/status/             — poll task progress
  3. GET    /parents/bulk-upload/template/                — download CSV/Excel template
  4. POST   /parents/bulk-upload/<id>/export-credentials/ — export login details
  5. GET    /parents/bulk-upload/<id>/errors/             — download error report CSV

All views are tenant-aware and admin-only.
Plain Django views (no @api_view) to avoid DRF auth dispatch issues.
"""

import io
import csv
import json
import logging
import os
import tempfile
from datetime import datetime

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt

from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


# ---------------------------------------------------------------------------
# Auth + tenant helpers
# ---------------------------------------------------------------------------

def _jwt_auth(request):
    """
    Validate JWT from Authorization header or httpOnly cookie.
    Returns the User instance on success, or None.
    """
    from django.conf import settings as django_settings

    # 1. Try Authorization: Bearer <token>
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if auth_header.startswith("Bearer "):
        try:
            jwt_auth = JWTAuthentication()
            validated = jwt_auth.get_validated_token(auth_header[7:])
            return jwt_auth.get_user(validated)
        except (InvalidToken, TokenError):
            pass

    # 2. Try httpOnly cookie fallback
    cookie_name = getattr(django_settings, "AUTH_COOKIE_ACCESS", "")
    if cookie_name:
        raw_token = request.COOKIES.get(cookie_name)
        if raw_token:
            try:
                jwt_auth = JWTAuthentication()
                validated = jwt_auth.get_validated_token(raw_token)
                return jwt_auth.get_user(validated)
            except (InvalidToken, TokenError):
                pass

    return None


def _get_tenant(request):
    return getattr(request, "tenant", None)


def _require_admin(request):
    """
    Returns (user, None) if authenticated admin, or (None, JsonResponse error).
    Usage:
        user, err = _require_admin(request)
        if err: return err
    """
    if request.method == "OPTIONS":
        return None, None  # handled by caller

    user = _jwt_auth(request)
    if not user:
        return None, JsonResponse({"error": "Authentication required."}, status=401)
    if not user.is_staff:
        return None, JsonResponse({"error": "Admin access required."}, status=403)
    return user, None


# ---------------------------------------------------------------------------
# 1. Upload endpoint
# ---------------------------------------------------------------------------

@csrf_exempt
def bulk_upload_parents(request):
    """
    POST /parents/bulk-upload/

    Accept a CSV or Excel file and queue it for async processing.

    Body (multipart/form-data):
      file  — required: the CSV or Excel file

    Returns immediately with { upload_id, task_id, status: 'pending' }.
    Frontend polls /bulk-upload/<upload_id>/status/ for progress.
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "POST, OPTIONS"
        return response

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user, err = _require_admin(request)
    if err:
        return err

    tenant = _get_tenant(request)
    if not tenant:
        return JsonResponse({"error": "Tenant not identified."}, status=400)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return JsonResponse({"error": "No file provided."}, status=400)

    _, ext = os.path.splitext(file_obj.name)
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return JsonResponse(
            {"error": f"Unsupported file type '{ext}'. Upload .csv, .xlsx, or .xls."},
            status=400,
        )

    # Save to a temp location the Celery worker can read
    upload_dir = os.path.join(tempfile.gettempdir(), "bulk_uploads")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"parent_bulk_{tenant.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    file_path = os.path.join(upload_dir, safe_name)
    with open(file_path, "wb") as dest:
        for chunk in file_obj.chunks():
            dest.write(chunk)

    # Create tracking record
    from parent.models import BulkUploadRecord
    record = BulkUploadRecord.objects.create(
        tenant=tenant,
        uploaded_by=user,
        original_filename=file_obj.name,
        file_path=file_path,
        file_ext=ext,
        status="pending",
    )

    # Enqueue Celery task
    from parent.tasks import process_bulk_parent_upload
    task = process_bulk_parent_upload.delay(
        upload_record_id=record.pk,
        tenant_id=str(tenant.id),
        file_path=file_path,
        file_ext=ext,
        uploaded_by_id=user.pk,
    )

    record.result_data = {"celery_task_id": task.id}
    record.save(update_fields=["result_data"])

    return JsonResponse(
        {
            "upload_id": record.pk,
            "task_id": task.id,
            "status": "pending",
            "message": "File accepted. Processing started.",
        },
        status=202,
    )


# ---------------------------------------------------------------------------
# 2. Status polling endpoint
# ---------------------------------------------------------------------------

@csrf_exempt
def bulk_upload_status(request, upload_id):
    """
    GET /parents/bulk-upload/<upload_id>/status/
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "GET, OPTIONS"
        return response

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user, err = _require_admin(request)
    if err:
        return err

    tenant = _get_tenant(request)

    from parent.models import BulkUploadRecord
    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        return JsonResponse({"error": "Upload record not found."}, status=404)

    payload = {
        "upload_id": record.pk,
        "status": record.status,
        "progress": record.progress_percent,
        "total_rows": record.total_rows,
        "processed_rows": record.processed_rows,
        "imported_rows": record.imported_rows,
        "failed_rows": record.failed_rows,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "original_filename": record.original_filename,
    }

    if record.status in ("completed", "failed"):
        payload["result"] = record.result_data

    return JsonResponse(payload)


# ---------------------------------------------------------------------------
# 3. Template download
# ---------------------------------------------------------------------------

@csrf_exempt
def download_upload_template(request):
    """
    GET /parents/bulk-upload/template/?format=csv|excel

    Download a blank CSV or Excel template with all required columns.
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "GET, OPTIONS"
        return response

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user, err = _require_admin(request)
    if err:
        return err

    fmt = request.GET.get("format", "csv").lower()

    # (header, required, example, description)
    COLUMNS = [
        ("Last Name*",            True,  "Adeyemi",            "Parent's surname / last name"),
        ("First Name*",           True,  "Fatima",             "Parent's first name"),
        ("Gender*",               True,  "F",                  "M or F"),
        ("Phone Number*",         True,  "08012345678",        "Parent's phone — used to link to students"),
        ("Email",                 False, "fatima@example.com", "Parent email (optional, auto-generated if blank)"),
        ("Address*",              True,  "12 Broad St, Lagos", "Home address"),
        ("Parent/Guardian Role*", True,  "Father",             "Father / Mother / Guardian / Sponsor"),
    ]

    headers = [col[0] for col in COLUMNS]
    example = [col[2] for col in COLUMNS]

    if fmt == "excel":
        return _template_excel(headers, example, COLUMNS)
    return _template_csv(headers, example)


def _template_csv(headers, example):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(example)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="parent_upload_template.csv"'
    return response


def _template_excel(headers, example, column_defs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _template_csv(headers, example)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parent Upload Template"

    REQUIRED_COLOR = "FFF2CC"
    OPTIONAL_COLOR = "FFFFFF"
    HEADER_COLOR   = "305496"
    HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
    EXAMPLE_FONT   = Font(italic=True, color="595959", size=10)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value="PARENT BULK UPLOAD TEMPLATE")
    title.font = Font(size=14, bold=True, color="1F4E79")
    title.alignment = Alignment(horizontal="center")

    # Instructions
    instructions = [
        "INSTRUCTIONS:",
        "• Do NOT delete the header row",
        "• Required fields (*) are highlighted in yellow",
        "• Gender must be M or F",
        "• Phone number is used to auto-link parents to existing students",
        "• Email is optional — one will be auto-generated if left blank",
        "• Upload PARENTS before uploading STUDENTS",
    ]
    for i, text in enumerate(instructions, start=2):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(size=10, color="444444")

    start_row = len(instructions) + 3

    # Legend
    ws.cell(row=start_row - 1, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=start_row - 1, column=2, value="Required").fill = PatternFill("solid", fgColor=REQUIRED_COLOR)
    ws.cell(row=start_row - 1, column=3, value="Optional").fill = PatternFill("solid", fgColor=OPTIONAL_COLOR)

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Example row
    for col_idx, val in enumerate(example, 1):
        col_def  = column_defs[col_idx - 1]
        required = col_def[1]
        cell = ws.cell(row=start_row + 1, column=col_idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = PatternFill("solid", fgColor=REQUIRED_COLOR if required else OPTIONAL_COLOR)
        cell.border = thin_border

    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[start_row].height = 30
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row}"

    # Field Guide sheet
    guide = wb.create_sheet("Field Guide")
    guide_headers = ["Column", "Required", "Example", "Description"]
    for col_idx, h in enumerate(guide_headers, 1):
        cell = guide.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.border = thin_border

    for row_idx, (header, required, ex, desc) in enumerate(column_defs, 2):
        guide.cell(row=row_idx, column=1, value=header).border = thin_border
        req_cell = guide.cell(row=row_idx, column=2, value="Yes" if required else "No")
        req_cell.font = Font(color="C00000" if required else "595959", bold=required)
        req_cell.border = thin_border
        guide.cell(row=row_idx, column=3, value=ex).border = thin_border
        desc_cell = guide.cell(row=row_idx, column=4, value=desc)
        desc_cell.alignment = Alignment(wrap_text=True)
        desc_cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="parent_upload_template.xlsx"'
    return response


# ---------------------------------------------------------------------------
# 4. Credential export
# ---------------------------------------------------------------------------

CRED_HEADERS = [
    "Full Name", "Phone", "Role",
    "Username", "Password (initial)", "Email",
]


def _cred_rows(imported):
    for entry in imported:
        yield [
            entry.get("full_name", ""),
            entry.get("phone", ""),
            entry.get("role", ""),
            entry.get("username", ""),
            entry.get("password", ""),
            entry.get("email", ""),
        ]


@csrf_exempt
def export_credentials(request, upload_id):
    """
    POST /parents/bulk-upload/<upload_id>/export-credentials/

    Body (JSON): { "format": "csv" | "excel" | "pdf" }
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "POST, OPTIONS"
        return response

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user, err = _require_admin(request)
    if err:
        return err

    tenant = _get_tenant(request)

    from parent.models import BulkUploadRecord
    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        return JsonResponse({"error": "Upload record not found."}, status=404)

    if record.status != "completed":
        return JsonResponse(
            {"error": "Upload is not yet complete. Cannot export credentials."},
            status=400,
        )

    imported = record.result_data.get("imported", [])
    if not imported:
        return JsonResponse({"error": "No successfully imported parents found."}, status=404)

    # Parse JSON body (frontend sends Content-Type: application/json)
    try:
        body = json.loads(request.body)
        fmt = body.get("format", "csv").lower()
    except (json.JSONDecodeError, AttributeError):
        fmt = request.POST.get("format", "csv").lower()

    if fmt not in ("csv", "excel", "pdf"):
        return JsonResponse({"error": "format must be csv, excel, or pdf."}, status=400)

    if fmt == "csv":
        return _export_csv(imported, record)
    elif fmt == "excel":
        return _export_excel(imported, record)
    else:
        return _export_pdf(imported, record)


def _export_csv(imported, record):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CRED_HEADERS)
    for row in _cred_rows(imported):
        writer.writerow(row)
    buf.seek(0)
    fname = f"parent_credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def _export_excel(imported, record):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _export_csv(imported, record)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parent Credentials"

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    PASS_FILL   = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(CRED_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(_cred_rows(imported), 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            if col_idx == 5:  # Password column
                cell.fill = PASS_FILL
                cell.font = Font(name="Courier New", size=10)

    for i, w in enumerate([28, 18, 15, 24, 24, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Upload ID";    summary["B1"] = record.id
    summary["A2"] = "File";         summary["B2"] = record.original_filename
    summary["A3"] = "Imported";     summary["B3"] = record.imported_rows
    summary["A4"] = "Skipped";      summary["B4"] = record.failed_rows
    summary["A5"] = "Export Date";  summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary["A7"] = "IMPORTANT"
    summary["A7"].font = Font(bold=True, color="C00000")
    summary["B7"] = "Passwords shown are initial. Parents should change them on first login."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"parent_credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def _export_pdf(imported, record):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        logger.warning("reportlab not installed — falling back to CSV export")
        return _export_csv(imported, record)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                 fontSize=18, spaceAfter=6, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("Sub", parent=styles["Normal"],
                                 fontSize=10, spaceAfter=4, alignment=TA_CENTER,
                                 textColor=colors.grey)
    warn_style  = ParagraphStyle("Warn", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#C00000"))

    elements = [
        Paragraph("Parent Login Credentials", title_style),
        Paragraph(
            f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  "
            f"File: {record.original_filename}  |  "
            f"Total imported: {record.imported_rows}",
            sub_style,
        ),
        Spacer(1, 4*mm),
        Paragraph(
            "⚠ Confidential — These are initial passwords. "
            "Distribute securely and advise parents to change on first login.",
            warn_style,
        ),
        Spacer(1, 6*mm),
    ]

    NAVY  = colors.HexColor("#1F4E79")
    AMBER = colors.HexColor("#FFF2CC")
    STRIPE = colors.HexColor("#F2F2F2")

    table_data = [CRED_HEADERS] + list(_cred_rows(imported))
    col_widths = [50*mm, 30*mm, 25*mm, 40*mm, 38*mm, 50*mm]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 9),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 8),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, STRIPE]),
        ("BACKGROUND",    (4, 1), (4, -1), AMBER),   # Password column
        ("FONTNAME",      (4, 1), (4, -1), "Courier"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)
    fname = f"parent_credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# 5. Error report download
# ---------------------------------------------------------------------------

@csrf_exempt
def download_error_report(request, upload_id):
    """
    GET /parents/bulk-upload/<upload_id>/errors/
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "GET, OPTIONS"
        return response

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user, err = _require_admin(request)
    if err:
        return err

    tenant = _get_tenant(request)

    from parent.models import BulkUploadRecord  # ← correct model
    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        return JsonResponse({"error": "Upload record not found."}, status=404)

    errors = record.result_data.get("errors", [])
    if not errors:
        return JsonResponse({"message": "No errors to report."}, status=200)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Row Number", "First Name", "Last Name", "Phone", "Role", "Error Details"])
    for err_entry in errors:
        data = err_entry.get("data", {})
        writer.writerow([
            err_entry.get("row", ""),
            data.get("first_name", ""),
            data.get("last_name", ""),
            data.get("phone", ""),
            data.get("relationship", ""),
            " | ".join(err_entry.get("errors", [])),
        ])

    buf.seek(0)
    fname = f"parent_upload_errors_{record.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response