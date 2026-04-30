"""
students/bulk_views.py

Views for:
  1. POST   /students/bulk-upload/           — accept file, enqueue Celery task
  2. GET    /students/bulk-upload/<id>/status/ — poll task progress
  3. GET    /students/bulk-upload/template/   — download pre-filled CSV/Excel template
  4. POST   /students/bulk-upload/<id>/export-credentials/
                                              — export login details (CSV / Excel / PDF)

All views are tenant-aware and admin-only.
"""

import io
import csv
import logging
import os
import tempfile
from datetime import datetime

from django.http import HttpResponse, FileResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from tenants.mixins import TenantFilterMixin
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

# ---------------------------------------------------------------------------
# Helper: get tenant from request
# ---------------------------------------------------------------------------

def _get_tenant(request):
    return getattr(request, "tenant", None)


# ---------------------------------------------------------------------------
# 1. Upload endpoint
# ---------------------------------------------------------------------------

@api_view(["POST"])
@permission_classes([IsAdminUser])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_students(request):
    """
    Accept a CSV or Excel file and queue it for async processing.

    Body (multipart/form-data):
      file               — required: the CSV or Excel file
      academic_session   — optional: PK of AcademicSession used to resolve Sections

    Returns immediately with { upload_id, task_id, status: 'pending' }.
    Frontend polls /bulk-upload/<upload_id>/status/ for progress.
    """
    tenant = _get_tenant(request)
    if not tenant:
        return Response({"error": "Tenant not identified."}, status=400)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"error": "No file provided."}, status=400)

    _, ext = os.path.splitext(file_obj.name)
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return Response(
            {"error": f"Unsupported file type '{ext}'. Upload .csv, .xlsx, or .xls."},
            status=400,
        )

    academic_session_id = request.data.get("academic_session") or None

    # Save to a temp location the Celery worker can read
    upload_dir = os.path.join(tempfile.gettempdir(), "bulk_uploads")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"bulk_{tenant.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    file_path = os.path.join(upload_dir, safe_name)
    with open(file_path, "wb") as dest:
        for chunk in file_obj.chunks():
            dest.write(chunk)

    # Create tracking record
    from students.models import BulkUploadRecord
    record = BulkUploadRecord.objects.create(
        tenant=tenant,
        uploaded_by=request.user,
        original_filename=file_obj.name,
        file_path=file_path,
        file_ext=ext,
        status="pending",
    )

    # Enqueue Celery task
    from students.tasks import process_bulk_student_upload
    task = process_bulk_student_upload.delay(
        upload_record_id=record.pk,
        tenant_id=tenant.id,
        file_path=file_path,
        file_ext=ext,
        academic_session_id=int(academic_session_id) if academic_session_id else None,
        uploaded_by_id=request.user.pk,
    )

    # record.result_data = {"celery_task_id": task.id}
    # record.save(update_fields=["result_data"])

    return Response(
        {
            "upload_id": record.pk,
            "task_id": task.id,
            "status": "pending",
            "message": "File accepted. Processing started.",
        },
        status=202,
    )


# ---------------------------------------------------------------------------
# 2. Status polling endpoint
# ---------------------------------------------------------------------------

@api_view(["GET"])
@permission_classes([IsAdminUser])
def bulk_upload_status(request, upload_id):
    """
    Poll the processing status of a bulk upload.

    Returns:
      status        — pending | processing | completed | failed
      progress      — 0-100
      total_rows, processed_rows, imported_rows, failed_rows
      result        — present only when status == 'completed'
        result.summary    { total, imported, skipped }
        result.errors     [ { row, errors, data } ]
        result.imported   [ { row, full_name, username, password, ... } ]
    """
    from students.models import BulkUploadRecord

    tenant = _get_tenant(request)
    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        return Response({"error": "Upload record not found."}, status=404)

    payload = {
        "upload_id": record.pk,
        "status": record.status,
        "progress": record.progress_percent,
        "total_rows": record.total_rows,
        "processed_rows": record.processed_rows,
        "imported_rows": record.imported_rows,
        "failed_rows": record.failed_rows,
        "created_at": record.created_at,
        "original_filename": record.original_filename,
    }

    if record.status in ("completed", "failed"):
        raw = record.result_data or {}
        payload["result"] = {
            "summary": raw.get("summary", {"total": 0, "imported": 0, "skipped": 0}),
            "errors": raw.get("errors", []),
            "imported": raw.get("imported", []),
        }
        if record.status == "failed":
            payload["error"] = raw.get("error", "Upload failed.")

    return Response(payload)


# ---------------------------------------------------------------------------
# 3. Template download
# ---------------------------------------------------------------------------


def _jwt_auth(request):
    """Returns (user, None) or raises JsonResponse."""
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")

    # Try Bearer token from header
    if auth_header.startswith("Bearer "):
        try:
            jwt_auth = JWTAuthentication()
            validated_token = jwt_auth.get_validated_token(auth_header[7:])
            return jwt_auth.get_user(validated_token)
        except (InvalidToken, TokenError):
            pass

    # Try cookie fallback
    from django.conf import settings as django_settings

    raw_token = request.COOKIES.get(getattr(django_settings, "AUTH_COOKIE_ACCESS", ""))
    if raw_token:
        try:
            jwt_auth = JWTAuthentication()
            validated_token = jwt_auth.get_validated_token(raw_token)
            return jwt_auth.get_user(validated_token)
        except (InvalidToken, TokenError):
            pass

    return None


@csrf_exempt
def download_upload_template(request):
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "GET, OPTIONS"
        return response

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user = _jwt_auth(request)
    if not user:
        return JsonResponse({"error": "Authentication required."}, status=401)
    if not user.is_staff:
        return JsonResponse({"error": "Admin access required."}, status=403)

    fmt = request.GET.get("format", "csv").lower()

    tenant = _get_tenant(request)

    # ---- Fetch live tenant data for smart examples & conditional columns ----
    classes = []
    sections = []
    streams = []
    enable_streaming = True
    school_name = ""

    if tenant:
        from classroom.models import Class, Section, Stream

        classes = list(
            Class.objects.filter(tenant=tenant, is_active=True)
            .order_by("order")
            .values_list("name", flat=True)
        )
        sections = list(
            Section.objects.filter(tenant=tenant, is_active=True)
            .values_list("name", flat=True)
            .distinct()
            .order_by("name")
        )
        streams = list(
            Stream.objects.filter(tenant=tenant)
            .values_list("name", flat=True)
            .order_by("name")
        )
        school_name = tenant.name

        try:
            tenant_settings = tenant.settings
            enable_streaming = tenant_settings.enable_streaming
        except Exception:
            pass

    # Use real tenant values as example data where available
    example_class = classes[0] if classes else "JSS 1"
    example_section = sections[0] if sections else "Gold"
    example_stream = streams[0] if streams else "SCIENCE"

    class_desc = (
        f"Select from dropdown — valid values: {', '.join(classes[:8])}"
        if classes
        else "Select from dropdown — must match your school"
    )
    section_desc = (
        f"Select from dropdown — valid values: {', '.join(sections[:8])}"
        if sections
        else "Select from dropdown — must match your school"
    )

    COLUMNS = [
        ("Registration Number", False, "0001", "School-assigned reg number"),
        ("Surname*", True, "Adeyemi", "Student's surname / last name"),
        ("First Name*", True, "Fatima", "Student's first name"),
        ("Middle Name", False, "Grace", "Optional middle name"),
        ("Gender*", True, "M", "M or F — select from dropdown"),
        ("Date of Birth*", True, "2010-03-15", "Format: YYYY-MM-DD"),
        ("State of Origin", False, "Lagos", "Select from dropdown — optional"),
        ("LGA", False, "Ikeja", "Local Government Area — optional"),
        ("Blood Group", False, "O+", "e.g. A+, B-, O+, AB+"),
        ("Place of Birth*", True, "Lagos", "City / town of birth"),
        ("Class Name*", True, example_class, class_desc),
        ("Section*", True, example_section, section_desc),
    ]

    if enable_streaming:
        stream_desc = (
            f"Required for Senior Secondary — valid values: {', '.join(streams[:6])}"
            if streams
            else "Required only for Senior Secondary — select from dropdown"
        )
        COLUMNS.append(("Stream", False, example_stream, stream_desc))

    COLUMNS += [
        ("Year Admitted*", True, "2024", "4-digit year e.g. 2024"),
        ("Admission Date*", True, "2024-09-02", "Format: YYYY-MM-DD"),
        ("Is Active", False, "TRUE", "TRUE or FALSE. Defaults to TRUE."),
        ("Address*", True, "12 Broad St, Lagos", "Home address"),
        ("Phone Number", False, "08012345678", "Student's phone"),
        (
            "Parent Contact*",
            True,
            "08098765432",
            "Parent/guardian phone — must already exist",
        ),
        ("Emergency Contact*", True, "08011112222", "Emergency contact number"),
        (
            "Parent/Guardian Name*",
            True,
            "Mr. Adeyemi Bola",
            "Full name of parent/guardian",
        ),
        (
            "Parent/Guardian Role*",
            True,
            "Father",
            "Father / Mother / Guardian / Sponsor",
        ),
        ("Medical Conditions", False, "Asthma", "Comma-separated list"),
        ("Special Requirements", False, "Wheelchair access", "Any special needs"),
        (
            "Profile Picture URL",
            False,
            "https://res.cloudinary.com/...",
            "Cloudinary URL",
        ),
        ("Email", False, "fatima@example.com", "Student email (optional)"),
    ]

    headers = [col[0] for col in COLUMNS]
    example = [col[2] for col in COLUMNS]

    if fmt == "excel":
        return _template_excel(
            headers,
            example,
            COLUMNS,
            tenant=tenant,
            school_name=school_name,
            classes=classes,
            sections=sections,
            streams=streams,
            enable_streaming=enable_streaming,
        )
    return _template_csv(headers, example)


def _template_csv(headers, example):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(example)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="student_upload_template.csv"'
    return response


def _template_excel(
    headers,
    example,
    column_defs,
    tenant=None,
    school_name="",
    classes=None,
    sections=None,
    streams=None,
    enable_streaming=True,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return _template_csv(headers, example)

    import io
    from django.http import HttpResponse

    # Use pre-fetched tenant data (already resolved by the view)
    classes = classes or []
    sections = sections or []
    streams = streams or []

    if not streams:
        streams = ["SCIENCE", "ARTS", "COMMERCIAL"]

    BLOOD_GROUP_CHOICES = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]

    NIGERIA_STATES = [
        "Abia",
        "Adamawa",
        "Akwa Ibom",
        "Anambra",
        "Bauchi",
        "Bayelsa",
        "Benue",
        "Borno",
        "Cross River",
        "Delta",
        "Ebonyi",
        "Edo",
        "Ekiti",
        "Enugu",
        "FCT",
        "Gombe",
        "Imo",
        "Jigawa",
        "Kaduna",
        "Kano",
        "Katsina",
        "Kebbi",
        "Kogi",
        "Kwara",
        "Lagos",
        "Nasarawa",
        "Niger",
        "Ogun",
        "Ondo",
        "Osun",
        "Oyo",
        "Plateau",
        "Rivers",
        "Sokoto",
        "Taraba",
        "Yobe",
        "Zamfara",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Upload Template"

    # ---- Hidden sheet for dropdown sources ----
    ref_ws = wb.create_sheet("_Ref")
    ref_ws.sheet_state = "hidden"

    def _write_ref_col(sheet, col, items):
        if not items:
            return None
        for i, val in enumerate(items, 1):
            sheet.cell(row=i, column=col, value=val)
        last = get_column_letter(col)
        return f"_Ref!${last}$1:${last}${len(items)}"

    gender_src = _write_ref_col(ref_ws, 1, ["M", "F"])
    class_src = _write_ref_col(ref_ws, 2, classes)
    section_src = _write_ref_col(ref_ws, 3, sections)
    stream_src = _write_ref_col(ref_ws, 4, streams)
    state_src = _write_ref_col(ref_ws, 5, NIGERIA_STATES)
    parent_role_src = _write_ref_col(
        ref_ws, 6, ["Father", "Mother", "Guardian", "Sponsor"]
    )
    blood_group_src = _write_ref_col(ref_ws, 7, BLOOD_GROUP_CHOICES)

    # ---- Styles ----
    REQUIRED_COLOR = "FFF2CC"
    OPTIONAL_COLOR = "FFFFFF"
    HEADER_COLOR = "305496"
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    EXAMPLE_FONT = Font(italic=True, color="595959", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ---- Title ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_text = (
        f"{school_name.upper()} — STUDENT BULK UPLOAD TEMPLATE"
        if school_name
        else "STUDENT BULK UPLOAD TEMPLATE"
    )
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # ---- Instructions ----
    instructions = [
        "INSTRUCTIONS:",
        "• Do NOT delete the header row",
        "• Required fields are highlighted in yellow",
        "• Dates must be YYYY-MM-DD format",
        "• Gender must be M or F — use the dropdown",
    ]

    if classes:
        instructions.append(
            f"• Valid Class Names: {', '.join(classes)} — use the dropdown"
        )
    else:
        instructions.append("• Class Name must match your school's setup — use the dropdown")

    if sections:
        instructions.append(
            f"• Valid Sections: {', '.join(sections)} — use the dropdown"
        )
    else:
        instructions.append("• Section name must match exactly (varies per school) — use the dropdown")

    if enable_streaming:
        if streams:
            instructions.append(
                f"• Stream (Senior Secondary only) — valid values: {', '.join(streams)}"
            )
        else:
            instructions.append("• Stream is only required for Senior Secondary — use the dropdown")

    instructions += [
        "• Parent phone must already exist in the system",
        "• State and LGA are optional — can be filled in later via student profile",
    ]

    for i, text in enumerate(instructions, start=2):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(size=10, color="444444")

    start_row = len(instructions) + 3

    # ---- Legend ----
    ws.cell(row=start_row - 1, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=start_row - 1, column=2, value="Required").fill = PatternFill(
        "solid", fgColor=REQUIRED_COLOR
    )
    ws.cell(row=start_row - 1, column=3, value="Optional").fill = PatternFill(
        "solid", fgColor=OPTIONAL_COLOR
    )

    # ---- Header row ----
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # ---- Example row ----
    for col_idx, val in enumerate(example, 1):
        col_def = column_defs[col_idx - 1] if col_idx - 1 < len(column_defs) else None
        required = col_def[1] if col_def else False
        cell = ws.cell(row=start_row + 1, column=col_idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = PatternFill(
            "solid", fgColor=REQUIRED_COLOR if required else OPTIONAL_COLOR
        )
        cell.border = thin_border

    # ---- Column widths ----
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[start_row].height = 30

    # ---- Data validation dropdowns ----
    # Map header name → validation source
    DROPDOWN_MAP = {
        "gender*": gender_src,
        "class name*": class_src,
        "section*": section_src,
        "stream": stream_src,
        "state of origin": state_src,
        "parent/guardian role*": parent_role_src,
        "blood group": blood_group_src,
    }

    DATA_ROWS = f"{start_row + 2}:{start_row + 1001}"  # up to 1000 data rows

    for col_idx, header in enumerate(headers, 1):
        key = header.strip().lower()
        src = DROPDOWN_MAP.get(key)
        if not src:
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=src,
            allow_blank=True,
            showDropDown=False,  # False = show the arrow
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Please select a value from the dropdown list.",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{col_letter}{start_row + 2}:{col_letter}{start_row + 1001}"

    # ---- Field Guide sheet ----
    guide = wb.create_sheet("Field Guide")
    guide_headers = ["Column", "Required", "Example", "Description"]
    for col_idx, h in enumerate(guide_headers, 1):
        cell = guide.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.border = thin_border

    for row_idx, (header, required, ex, desc) in enumerate(column_defs, 2):
        guide.cell(row=row_idx, column=1, value=header).border = thin_border
        req_cell = guide.cell(row=row_idx, column=2, value="Yes" if required else "No")
        req_cell.font = Font(color="C00000" if required else "595959", bold=required)
        req_cell.border = thin_border
        guide.cell(row=row_idx, column=3, value=ex).border = thin_border
        desc_cell = guide.cell(row=row_idx, column=4, value=desc)
        desc_cell.alignment = Alignment(wrap_text=True)
        desc_cell.border = thin_border

    # ---- Save ----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_upload_template.xlsx"'
    return response


# ---------------------------------------------------------------------------
# 4. Credential export
# ---------------------------------------------------------------------------


@csrf_exempt
def export_credentials(request, upload_id):
    """
    Export login credentials for all successfully imported students
    from a completed bulk upload.

    Body:
      format — 'csv' | 'excel' | 'pdf'
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "POST, OPTIONS"
        return response

    if request.method != "POST":
        # Bug 1: used DRF Response instead of JsonResponse
        return JsonResponse({"error": "Method not allowed."}, status=405)

    user = _jwt_auth(request)
    if not user:
        return JsonResponse({"error": "Authentication required."}, status=401)
    if not user.is_staff:
        return JsonResponse({"error": "Admin access required."}, status=403)

    tenant = _get_tenant(request)

    from students.models import BulkUploadRecord

    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        return JsonResponse({"error": "Upload record not found."}, status=404)

    if record.status != "completed":
        return JsonResponse(
            {"error": "Upload is not yet complete. Cannot export credentials."},
            status=400,
        )

    imported = record.result_data.get("imported", [])
    if not imported:
        return JsonResponse(
            {"error": "No successfully imported students found."}, status=404
        )

    import json

    try:
        body = json.loads(request.body)
        fmt = body.get("format", "csv").lower()
    except (json.JSONDecodeError, AttributeError):
        fmt = request.POST.get("format", "csv").lower()

    if fmt not in ("csv", "excel", "pdf"):
        return JsonResponse({"error": "format must be csv, excel, or pdf."}, status=400)

    if fmt == "csv":
        return _export_csv(imported, record)
    elif fmt == "excel":
        return _export_excel(imported, record)
    else:
        return _export_pdf(imported, record)


def _cred_rows(imported):
    """Yield rows for credential tables."""
    for entry in imported:
        yield [
            entry.get("registration_number") or "—",
            entry.get("full_name", ""),
            entry.get("classroom", ""),
            entry.get("username", ""),
            entry.get("password", ""),
            entry.get("parent_name", ""),
            entry.get("parent_phone", ""),
        ]


CRED_HEADERS = [
    "Reg Number", "Full Name", "Classroom",
    "Username", "Password (initial)",
    "Parent Name", "Parent Phone",
]


def _export_csv(imported, record):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CRED_HEADERS)
    for row in _cred_rows(imported):
        writer.writerow(row)
    buf.seek(0)
    fname = f"credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def _export_excel(imported, record):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _export_csv(imported, record)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Credentials"

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    PASS_FILL    = PatternFill("solid", fgColor="FFF2CC")
    thin         = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Header
    for col_idx, h in enumerate(CRED_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, row_data in enumerate(_cred_rows(imported), 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            # Highlight password column
            if col_idx == 5:
                cell.fill = PASS_FILL
                cell.font = Font(name="Courier New", size=10)

    # Column widths
    for i, w in enumerate([18, 28, 20, 22, 22, 28, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Upload ID"
    summary["B1"] = record.id
    summary["A2"] = "File"
    summary["B2"] = record.original_filename
    summary["A3"] = "Imported"
    summary["B3"] = record.imported_rows
    summary["A4"] = "Skipped"
    summary["B4"] = record.failed_rows
    summary["A5"] = "Export Date"
    summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary["A6"] = ""
    summary["A7"] = "IMPORTANT"
    summary["A7"].font = Font(bold=True, color="C00000")
    summary["B7"] = "Passwords shown are initial. Students/parents should change them on first login."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def _export_pdf(imported, record):
    """
    Generate a PDF with:
      - Cover page summary
      - Table of all credentials
    Uses reportlab. Falls back to CSV if not installed.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, PageBreak,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        logger.warning("reportlab not installed — falling back to CSV export")
        return _export_csv(imported, record)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, spaceAfter=6, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=10, spaceAfter=4, alignment=TA_CENTER, textColor=colors.grey,
    )
    warn_style = ParagraphStyle(
        "Warn", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#C00000"),
    )

    elements = []

    # Cover / header
    elements.append(Paragraph("Student Login Credentials", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  "
        f"Upload: {record.original_filename}  |  "
        f"Total imported: {record.imported_rows}",
        sub_style,
    ))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        "⚠ Confidential — These are initial passwords. "
        "Distribute securely and advise students/parents to change on first login.",
        warn_style,
    ))
    elements.append(Spacer(1, 6 * mm))

    # Table
    NAVY  = colors.HexColor("#1F4E79")
    AMBER = colors.HexColor("#FFF2CC")
    STRIPE = colors.HexColor("#F2F2F2")

    table_data = [CRED_HEADERS]
    for row in _cred_rows(imported):
        table_data.append(row)

    col_widths = [30*mm, 50*mm, 35*mm, 40*mm, 38*mm, 50*mm, 30*mm]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND",   (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 9),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 8),
        ("TOPPADDING",   (0, 0), (-1, 0), 8),
        # Body
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("TOPPADDING",   (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, STRIPE]),
        # Password column highlight
        ("BACKGROUND",   (4, 1), (4, -1), AMBER),
        ("FONTNAME",     (4, 1), (4, -1), "Courier"),
        # Grid
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)
    fname = f"credentials_{record.id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# 5. Error report download
# ---------------------------------------------------------------------------


@csrf_exempt
def download_error_report(request, upload_id):
    """
    Download a CSV of all rows that failed validation for a completed upload.
    """
    if request.method == "OPTIONS":
        response = HttpResponse()
        response["Allow"] = "GET, OPTIONS"
        return response

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed."}, status=405)

    # Bug 1: _jwt_auth returns a USER not a tenant — variable was misnamed
    user = _jwt_auth(request)
    if not user:
        return JsonResponse({"error": "Authentication required."}, status=401)
    if not user.is_staff:
        return JsonResponse({"error": "Admin access required."}, status=403)

    # Bug 2: tenant must come from request, not from _jwt_auth
    tenant = _get_tenant(request)

    from students.models import BulkUploadRecord

    try:
        record = BulkUploadRecord.objects.get(pk=upload_id, tenant=tenant)
    except BulkUploadRecord.DoesNotExist:
        # Bug 3: can't use DRF Response in a plain Django view — use JsonResponse
        return JsonResponse({"error": "Upload record not found."}, status=404)

    errors = record.result_data.get("errors", [])
    if not errors:
        # Bug 3: same — use JsonResponse not Response
        return JsonResponse({"message": "No errors to report."}, status=200)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Row Number", "First Name", "Last Name", "Reg Number", "Class", "Section", "Error Details"])
    for err in errors:
        data = err.get("data", {})
        writer.writerow([
            err.get("row", ""),
            data.get("first_name", ""),
            data.get("last_name", ""),
            data.get("registration_number", ""),
            data.get("class_code", ""),
            data.get("section_name", ""),
            " | ".join(err.get("errors", [])),
        ])

    buf.seek(0)
    fname = f"upload_errors_{record.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(buf.read(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response
